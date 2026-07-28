# -*- coding: utf-8 -*-
"""Build app-ready JSON from the Korea District Heating Corp. Excel files."""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
ROOT_DIR = PROJECT_DIR.parent
PUBLIC_DATA_DIR = PROJECT_DIR / "public" / "data"
OUTPUT_PATH = PUBLIC_DATA_DIR / "hanan-datasets.json"


FILE_HEADERS = [
    "순번",
    "파일아이디",
    "검색일자",
    "파일데이터명",
    "분류체계",
    "제공기관",
    "관리부서명",
    "관리부서전화번호",
    "보유근거",
    "수집방법",
    "업데이트주기",
    "등록일",
    "수정일",
    "차기등록예정일",
    "매체유형",
    "전체행",
    "확장자",
    "조회수",
    "누적다운로드수",
    "다운로드수",
    "데이터한계",
    "키워드",
    "제공형태",
    "설명",
    "기타유의사항",
    "URL",
    "최초등록일",
]

API_HEADERS = [
    "순번",
    "API_ID",
    "검색일자",
    "API_명",
    "분류체계",
    "제공기관",
    "관리부서명",
    "관리부서 전화번호",
    "API 유형",
    "데이터포맷",
    "조회건수",
    "활용신청",
    "키워드",
    "등록일",
    "수정일",
    "비용부과유무",
    "신청가능 트래픽",
    "심의유형",
    "이용허락범위",
    "참고문서",
    "API_URL",
]

THEME_RULES = [
    (
        "AI·이미지",
        ["학습용", "이미지", "사진", "동영상", "영상", "AI", "JPG", "MP4", "AVI"],
    ),
    (
        "에너지·열공급",
        ["열공급", "연료", "발전", "전력", "열수요", "열생산", "판매량", "사용량", "난방지수", "열량", "열원", "에너지"],
    ),
    (
        "설비·자산",
        ["설비", "배관", "열수송", "맨홀", "밸브", "보일러", "기계실", "시설", "자산", "준공", "공급시설", "기자재", "정비"],
    ),
    (
        "환경·안전",
        ["환경", "안전", "온실가스", "배출", "기상", "태양광", "태양열", "수질", "재난", "사고", "위험", "보호구", "ISO", "폐기물"],
    ),
    (
        "지역·공간",
        ["지역별", "지사별", "권역", "주소", "위치", "지도", "관망", "공간", "광주전남", "판교", "지역난방"],
    ),
    (
        "고객·요금",
        ["고객", "요금", "민원", "계량기", "검침", "사용자", "공동주택", "세대", "계약종별"],
    ),
    (
        "경영·행정",
        ["입찰", "계약", "경영", "예산", "조직", "인사", "채용", "ESG", "공시", "감사", "제도", "보도자료", "교육", "연구"],
    ),
]


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value: Any) -> int:
    value_text = text(value).replace(",", "")
    if not value_text:
        return 0
    try:
        return int(float(value_text))
    except ValueError:
        return 0


def iso_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    value_text = text(value)
    if re.fullmatch(r"\d{8}", value_text):
        return f"{value_text[:4]}-{value_text[4:6]}-{value_text[6:]}"
    return value_text


def clean_title(value: Any) -> str:
    title = text(value)
    title = re.sub(r"^한국지역난방공사_?", "", title)
    title = re.sub(r"_?\d{8}$", "", title)
    title = re.sub(r"_?\d{6}$", "", title)
    title = re.sub(r"\s+", " ", title)
    return title.strip(" _")


def slugify(value: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z가-힣]+", "-", value).strip("-")
    return slug[:90] or "dataset"


def split_keywords(value: Any) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for token in re.split(r"[,;/|\n\r]+", text(value)):
        keyword = re.sub(r"\s+", " ", token).strip()
        if len(keyword) > 1 and keyword not in seen:
            result.append(keyword)
            seen.add(keyword)
    return result


def compact(value: Any, limit: int = 360) -> str:
    content = re.sub(r"\s+", " ", text(value)).strip()
    if len(content) <= limit:
        return content
    return content[:limit].rstrip() + "..."


def category_group(category: str) -> str:
    return category.split(" - ", 1)[0].strip() if " - " in category else category


def format_phone(value: Any) -> str:
    raw = text(value)
    if not raw:
        return ""
    digits = re.sub(r"\D+", "", raw.split(".", 1)[0])
    if len(digits) == 10 and not digits.startswith("0"):
        digits = "0" + digits
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return raw


def theme_for(record: dict[str, Any]) -> str:
    primary_blob = " ".join(
        [
            record.get("name", ""),
            record.get("format", ""),
            " ".join(record.get("keywords", [])),
        ]
    )
    for theme, tokens in THEME_RULES:
        if any(token in primary_blob for token in tokens):
            return theme

    category = record.get("category", "")
    if "환경" in category or "공공질서및안전" in category:
        return "환경·안전"
    if "사회복지" in category:
        return "고객·요금"
    if any(token in category for token in ["일반공공행정", "과학기술", "교육", "문화체육관광"]):
        return "경영·행정"
    if "산업·통상·중소기업" in category:
        return "에너지·열공급"
    return "기타"


def row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}


def locate_workbooks() -> tuple[Path, Path, Path]:
    candidates: list[tuple[Path, int, int]] = []
    for path in sorted(ROOT_DIR.glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        candidates.append((path, len(wb.sheetnames), ws.max_column))

    file_book = next(path for path, sheet_count, cols in candidates if sheet_count == 1 and cols == 27)
    api_book = next(path for path, sheet_count, cols in candidates if sheet_count == 1 and cols == 21)
    summary_book = next(path for path, sheet_count, _ in candidates if sheet_count == 3)
    return file_book, api_book, summary_book


def file_record(values: tuple[Any, ...], index: int) -> dict[str, Any]:
    row = row_dict(FILE_HEADERS, values)
    name = clean_title(row["파일데이터명"])
    category = text(row["분류체계"])
    record = {
        "id": f"file-{text(row['파일아이디']) or index}",
        "sourceId": text(row["파일아이디"]),
        "kind": "file",
        "name": name,
        "originalName": text(row["파일데이터명"]),
        "slug": slugify(name),
        "category": category,
        "categoryGroup": category_group(category),
        "provider": text(row["제공기관"]),
        "department": text(row["관리부서명"]),
        "departmentPhone": format_phone(row["관리부서전화번호"]),
        "searchDate": iso_date(row["검색일자"]),
        "createdAt": iso_date(row["등록일"]),
        "updatedAt": iso_date(row["수정일"]),
        "nextUpdateAt": iso_date(row["차기등록예정일"]),
        "firstRegisteredAt": iso_date(row["최초등록일"]),
        "updateCycle": text(row["업데이트주기"]),
        "mediaType": text(row["매체유형"]),
        "format": text(row["확장자"]) or "파일",
        "rowCount": number(row["전체행"]),
        "views": number(row["조회수"]),
        "downloads": number(row["다운로드수"]),
        "cumulativeDownloads": number(row["누적다운로드수"]),
        "applications": 0,
        "keywords": split_keywords(row["키워드"]),
        "provisionType": text(row["제공형태"]),
        "description": compact(row["설명"]),
        "limitations": compact(row["데이터한계"]),
        "notes": compact(row["기타유의사항"]),
        "url": text(row["URL"]),
        "legalBasis": compact(row["보유근거"]),
        "collectionMethod": compact(row["수집방법"]),
    }
    record["theme"] = theme_for(record)
    return record


def api_record(values: tuple[Any, ...], index: int) -> dict[str, Any]:
    row = row_dict(API_HEADERS, values)
    name = clean_title(row["API_명"])
    category = text(row["분류체계"])
    data_format = text(row["데이터포맷"])
    api_type = text(row["API 유형"])
    record = {
        "id": f"api-{text(row['API_ID']) or index}",
        "sourceId": text(row["API_ID"]),
        "kind": "api",
        "name": name,
        "originalName": text(row["API_명"]),
        "slug": slugify(name),
        "category": category,
        "categoryGroup": category_group(category),
        "provider": text(row["제공기관"]),
        "department": text(row["관리부서명"]),
        "departmentPhone": format_phone(row["관리부서 전화번호"]),
        "searchDate": iso_date(row["검색일자"]),
        "createdAt": iso_date(row["등록일"]),
        "updatedAt": iso_date(row["수정일"]),
        "nextUpdateAt": "",
        "firstRegisteredAt": "",
        "updateCycle": "API",
        "mediaType": "API",
        "format": "/".join(part for part in [api_type, data_format] if part) or "API",
        "rowCount": 0,
        "views": number(row["조회건수"]),
        "downloads": 0,
        "cumulativeDownloads": 0,
        "applications": number(row["활용신청"]),
        "keywords": split_keywords(row["키워드"]),
        "provisionType": data_format,
        "description": "",
        "limitations": "",
        "notes": "",
        "url": text(row["API_URL"]),
        "apiType": api_type,
        "dataFormat": data_format,
        "isCharged": text(row["비용부과유무"]),
        "traffic": text(row["신청가능 트래픽"]),
        "reviewType": text(row["심의유형"]),
        "license": text(row["이용허락범위"]),
        "referenceDocument": text(row["참고문서"]),
    }
    record["theme"] = theme_for(record)
    return record


def non_empty_rows(path: Path, headers: list[str]) -> list[tuple[Any, ...]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if any(value not in (None, "") for value in values):
            rows.append(tuple(values[: len(headers)]))
    return rows


def extract_catalog_summary(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in workbook.worksheets:
        source_note = text(ws.cell(2, 1).value)
        headers = [text(ws.cell(3, column).value) for column in range(1, ws.max_column + 1)]
        rows = []
        for excel_row in ws.iter_rows(min_row=4, values_only=True):
            if not any(value not in (None, "") for value in excel_row):
                continue
            item = {
                headers[index] or f"column{index + 1}": excel_row[index]
                for index in range(min(len(headers), len(excel_row)))
            }
            rows.append({key: text(value) if not isinstance(value, (int, float)) else value for key, value in item.items()})
        sheets.append(
            {
                "sheetName": ws.title,
                "title": text(ws.cell(1, 1).value),
                "source": source_note,
                "records": rows,
            }
        )
    return sheets


def counter_list(counter: Counter[str]) -> list[dict[str, Any]]:
    return [{"name": name, "count": count} for name, count in counter.most_common()]


def build_catalog() -> dict[str, Any]:
    file_book, api_book, summary_book = locate_workbooks()
    records = [
        file_record(row, index)
        for index, row in enumerate(non_empty_rows(file_book, FILE_HEADERS), start=1)
    ]
    records.extend(
        api_record(row, index)
        for index, row in enumerate(non_empty_rows(api_book, API_HEADERS), start=1)
    )
    records.sort(key=lambda item: (item["theme"], item["kind"], item["name"], item["sourceId"]))

    theme_counts = Counter(record["theme"] for record in records)
    category_counts = Counter(record["categoryGroup"] for record in records)
    format_counts = Counter(record["format"] for record in records)
    keyword_counts = Counter(keyword for record in records for keyword in record["keywords"])
    kind_counts = Counter(record["kind"] for record in records)
    search_dates = sorted({record["searchDate"] for record in records if record["searchDate"]})

    return {
        "source": {
            "organization": "한국지역난방공사",
            "portal": "공공데이터포털(data.go.kr)",
            "asOf": search_dates[-1] if search_dates else "",
            "generatedAt": date.today().isoformat(),
            "workbooks": [
                {"role": "fileData", "fileName": file_book.name},
                {"role": "openApi", "fileName": api_book.name},
                {"role": "catalogSummary", "fileName": summary_book.name},
            ],
        },
        "summary": {
            "total": len(records),
            "files": kind_counts.get("file", 0),
            "apis": kind_counts.get("api", 0),
            "views": sum(record["views"] for record in records),
            "downloads": sum(record["downloads"] for record in records),
            "cumulativeDownloads": sum(record["cumulativeDownloads"] for record in records),
            "applications": sum(record["applications"] for record in records),
            "byTheme": counter_list(theme_counts),
            "byCategoryGroup": counter_list(category_counts),
            "byFormat": counter_list(format_counts),
            "topKeywords": counter_list(keyword_counts)[:20],
        },
        "datasets": records,
        "catalogSummary": extract_catalog_summary(summary_book),
    }


def main() -> None:
    PUBLIC_DATA_DIR.mkdir(parents=True, exist_ok=True)
    catalog = build_catalog()
    OUTPUT_PATH.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(catalog['datasets'])} datasets to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
